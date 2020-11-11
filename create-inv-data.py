from openpyxl import load_workbook
import random
from datetime import datetime
import json


wb2 = load_workbook('sample.xlsx')
worksheet1 = wb2['Sheet1'] # one way to load a worksheet

items = []

for row in worksheet1.iter_rows():
    if row[3].value is not None:
        items.append(row[3].value)


items.remove('Item')    
items = set(items)
#print(items)

data= {}

for i in items:
    data[i]=(random.randint(5,15),datetime.timestamp(datetime.now()))
    
print(data)

with open('inv-data.json', 'w') as fp:
    json.dump(data, fp,indent = 4, sort_keys=True)