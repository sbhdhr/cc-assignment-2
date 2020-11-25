from openpyxl import load_workbook
import json

def getOperations():
    wb2 = load_workbook('sample.xlsx')
    worksheet1 = wb2['Sheet1'] # one way to load a worksheet
    items = []

    clients = [{},{},{},{},{}]

    for index,row in enumerate(worksheet1.iter_rows()):
        if(index==0):
            continue
        user = row[0].value
        
        node = row[1].value
        node = int(node[1:])

        operation = row[2].value

        if(user[0]=='U'):
            user = int(user[1:])
        else:
            user = 0

        item = row[3].value

        tmp = {"UserID":user,"Operation":operation,"item":item}
        clients[node-1].update({index:tmp})

        with open('client0.json', 'w') as fp:
            json.dump(clients[0], fp,indent = 4, sort_keys=True)
        

        with open('client1.json', 'w') as fp:
            json.dump(clients[1], fp,indent = 4, sort_keys=True)

        with open('client2.json', 'w') as fp:
            json.dump(clients[2], fp,indent = 4, sort_keys=True)
        
        with open('client3.json', 'w') as fp:
            json.dump(clients[3], fp,indent = 4, sort_keys=True)
        
        with open('client4.json', 'w') as fp:
            json.dump(clients[4], fp,indent = 4, sort_keys=True)
    

getOperations()
        