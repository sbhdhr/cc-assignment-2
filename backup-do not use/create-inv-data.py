from openpyxl import load_workbook
import random
import time 
import json
import simplejson

import hashlib
from crush import Crush

crushmap = """
{
   "trees":[
      {
         "type":"host",
         "name":"InventoryDataReplicas",
         "id":-2,
         "children":[
            {
               "id":0,
               "name":"invdata0",
               "weight":65536
            },
            {
               "id":1,
               "name":"invdata1",
               "weight":65536
            }
         ]
      },
      {
         "type":"host",
         "name":"CartDataReplicas",
         "id":-3,
         "children":[
            {
               "id":2,
               "name":"cartdata0",
               "weight":65536
            },
            {
               "id":3,
               "name":"cartdata1",
               "weight":65536
            }
         ]
      }
   ],
   "rules":{
      "InvRule":[
         [
            "take",
            "InventoryDataReplicas"
         ],
         [
            "chooseleaf",
            "firstn",
            0,
            "type",
            "host"
         ],
         [
            "emit"
         ]
      ],
      "CartRule":[
         [
            "take",
            "CartDataReplicas"
         ],
         [
            "chooseleaf",
            "firstn",
            0,
            "type",
            "host"
         ],
         [
            "emit"
         ]
      ]
   }
}
"""

crush = Crush()
crush.parse(simplejson.loads(crushmap))

def find_part_for_inv(item):
    #crush code
    itemToBeSent = int(hashlib.sha1(item).hexdigest(), 16) % (10 ** 8)
    arr = crush.map(rule="InvRule", value=itemToBeSent, replication_count=1)
    val = arr[0][-1:]
    return int(val)


wb2 = load_workbook('sample.xlsx')
worksheet1 = wb2['Sheet1'] # one way to load a worksheet

items = []

for row in worksheet1.iter_rows():
    if row[3].value is not None:
        items.append(row[3].value)


items.remove('Item')    
items = set(items)
#print(items)

data0= {}
data1={}

for i in items:
    item_dict={'quan':random.randint(15,30),'price':round(random.uniform(50, 250),2),'time':time.time()}
    part = find_part_for_inv(i)
    print "{} -> {}".format(i,part)
    if(part==0):
        data0[i]=item_dict
    else:
        data1[i]=item_dict


with open('inv-data-0.json', 'w') as fp:
    json.dump(data0, fp,indent = 4, sort_keys=True)
    

with open('inv-data-1.json', 'w') as fp:
    json.dump(data1, fp,indent = 4, sort_keys=True)
    